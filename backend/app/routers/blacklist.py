"""
Router Blacklist — CRUD de CPF, CNPJ, telefone e e-mail bloqueados.
"""

import csv
import io
import re
from datetime import datetime, timezone
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.database import SessionLocal
from app.models import Blacklist, TipoBlacklist, Usuario
from app.routers.auth import verificar_token
from app.services.auditoria import log_auditoria

router = APIRouter(prefix="/blacklist", tags=["blacklist"])


# ── Dependência de DB ─────────────────────────────────────────────────────────

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


# ── Helpers ───────────────────────────────────────────────────────────────────

def _normalizar(valor: str, tipo: str) -> str:
    v = valor.strip()
    if tipo in ("CPF", "CNPJ", "TELEFONE"):
        v = re.sub(r"\D", "", v)
    return v


def _to_dict(e: Blacklist) -> dict[str, Any]:
    return {
        "id": e.id,
        "tipo": e.tipo,
        "valor": e.valor,
        "motivo": e.motivo,
        "fonte": e.fonte,
        "adicionado_por": e.adicionado_por,
        "ativo": e.ativo,
        "criado_em": e.criado_em.isoformat() if e.criado_em else None,
        "atualizado_em": e.atualizado_em.isoformat() if e.atualizado_em else None,
    }


# ── Schemas ───────────────────────────────────────────────────────────────────

class EntradaBody(BaseModel):
    tipo: str
    valor: str
    motivo: str
    fonte: str | None = None
    adicionado_por: str | None = None


class CheckBody(BaseModel):
    tipo: str
    valor: str


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.get("/")
def listar(
    pagina: int = Query(1, ge=1),
    limite: int = Query(20, ge=1, le=200),
    tipo: str | None = None,
    busca: str | None = None,
    db: Session = Depends(get_db),
):
    q = db.query(Blacklist).filter(Blacklist.ativo == True)  # noqa: E712
    if tipo:
        q = q.filter(Blacklist.tipo == tipo.upper())
    if busca:
        q = q.filter(Blacklist.valor.ilike(f"%{busca}%"))
    total = q.count()
    items = q.order_by(Blacklist.criado_em.desc()).offset((pagina - 1) * limite).limit(limite).all()

    # Contagem por tipo sobre a base inteira (ativo=True), não sobre a página
    # atual — os cards da tela usam isso para não mostrar só o tamanho da
    # página (bug corrigido: ver M13 da auditoria).
    contagem_bruta = dict(
        db.query(Blacklist.tipo, func.count(Blacklist.id))
        .filter(Blacklist.ativo == True)  # noqa: E712
        .group_by(Blacklist.tipo)
        .all()
    )
    contagem_por_tipo = {t.value: contagem_bruta.get(t, 0) for t in TipoBlacklist}

    return {
        "total": total,
        "pagina": pagina,
        "items": [_to_dict(i) for i in items],
        "contagem_por_tipo": contagem_por_tipo,
    }


# ── Exportação Excel ──────────────────────────────────────────────────────────
# Endpoint dedicado — não reutiliza o exportador genérico de relatorios.py.

_COLUNAS_EXCEL = ["Tipo", "Valor", "Motivo", "Fonte", "Adicionado Por", "Ativo", "Criado Em", "Atualizado Em"]
_LARGURAS_EXCEL = [12, 20, 45, 22, 22, 10, 18, 18]


def _fmt_data_excel(dt) -> str:
    return dt.strftime("%d/%m/%Y %H:%M") if dt else ""


@router.get("/exportar-excel")
def exportar_excel(
    request: Request,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(verificar_token),
):
    """
    Exporta TODA a blacklist (sem paginação, sem limite) para .xlsx, ordenada
    por criado_em desc. Mesma permissão de leitura da blacklist — só exige
    estar autenticado, para poder registrar quem exportou na auditoria.
    """
    entradas = db.query(Blacklist).order_by(Blacklist.criado_em.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Blacklist"

    ws.append(_COLUNAS_EXCEL)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    for col_idx in range(1, len(_COLUNAS_EXCEL) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for e in entradas:
        ws.append([
            e.tipo.value if hasattr(e.tipo, "value") else e.tipo,
            e.valor,
            e.motivo,
            e.fonte or "",
            e.adicionado_por or "",
            "Sim" if e.ativo else "Não",
            _fmt_data_excel(e.criado_em),
            _fmt_data_excel(e.atualizado_em),
        ])

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    for col_idx, largura in enumerate(_LARGURAS_EXCEL, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = largura

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    total = len(entradas)
    nome_arquivo = f"blacklist_{datetime.now(timezone.utc).strftime('%Y-%m-%d_%H-%M')}.xlsx"

    log_auditoria(
        db,
        acao=f"BLACKLIST_EXPORTADA: exportou {total} registro(s) da blacklist para Excel",
        usuario=usuario,
        request=request,
        tipo_entidade="blacklist",
        entidade_id=None,
        depois={"evento": "BLACKLIST_EXPORTADA", "quantidade": total, "arquivo": nome_arquivo},
        risco="BAIXO",
    )
    db.commit()

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome_arquivo}"'},
    )


@router.post("/check")
def verificar(body: CheckBody, db: Session = Depends(get_db)):
    valor = _normalizar(body.valor, body.tipo.upper())
    entry = db.query(Blacklist).filter(
        Blacklist.tipo == body.tipo.upper(),
        Blacklist.valor == valor,
        Blacklist.ativo == True,  # noqa: E712
    ).first()
    return {"bloqueado": entry is not None, "motivo": entry.motivo if entry else None}


@router.post("/", status_code=201)
def criar(body: EntradaBody, db: Session = Depends(get_db)):
    tipo = body.tipo.upper()
    valor = _normalizar(body.valor, tipo)
    if not valor:
        raise HTTPException(status_code=422, detail="Valor não pode ser vazio.")
    existing = db.query(Blacklist).filter(Blacklist.tipo == tipo, Blacklist.valor == valor).first()
    if existing:
        if not existing.ativo:
            existing.ativo = True
            existing.motivo = body.motivo
            existing.fonte = body.fonte
            db.commit()
            db.refresh(existing)
            return _to_dict(existing)
        raise HTTPException(status_code=409, detail=f"{tipo} já está na blacklist.")
    entry = Blacklist(
        tipo=tipo,
        valor=valor,
        motivo=body.motivo,
        fonte=body.fonte,
        adicionado_por=body.adicionado_por,
        ativo=True,
    )
    db.add(entry)
    db.commit()
    db.refresh(entry)
    return _to_dict(entry)


@router.delete("/{entry_id}")
def remover(
    entry_id: str,
    db: Session = Depends(get_db),
    _: Usuario = Depends(verificar_token),
):
    entry = db.query(Blacklist).filter(Blacklist.id == entry_id).first()
    if not entry:
        raise HTTPException(status_code=404, detail="Entrada não encontrada.")
    db.delete(entry)
    db.commit()
    return {"mensagem": "Removido com sucesso."}


@router.post("/import")
async def importar(
    request: Request,
    arquivo: UploadFile = File(...),
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(verificar_token),
):
    if not arquivo.filename:
        raise HTTPException(status_code=422, detail="Arquivo inválido.")
    conteudo = await arquivo.read()
    try:
        texto = conteudo.decode("utf-8-sig")
    except UnicodeDecodeError:
        texto = conteudo.decode("latin-1")

    inseridos = 0
    pulados = 0
    erros: list[dict] = []
    reader = csv.DictReader(io.StringIO(texto))
    for i, row in enumerate(reader, start=2):
        try:
            with db.begin_nested():  # SAVEPOINT — isola falha desta linha sem perder as anteriores
                tipo_raw = (row.get("tipo") or "CPF").strip().upper()
                if tipo_raw not in ("CPF", "CNPJ", "TELEFONE", "EMAIL"):
                    tipo_raw = "CPF"
                valor_raw = (
                    row.get("valor") or row.get("cpf") or row.get("cnpj") or
                    row.get("telefone") or row.get("email") or ""
                ).strip()
                motivo = (row.get("motivo") or "Importação em lote").strip()
                fonte = (row.get("fonte") or "importação").strip()
                if not valor_raw:
                    pulados += 1
                    continue
                valor = _normalizar(valor_raw, tipo_raw)
                if not valor:
                    raise ValueError("Valor vazio após normalização")
                existing = db.query(Blacklist).filter(
                    Blacklist.tipo == tipo_raw, Blacklist.valor == valor
                ).first()
                if existing:
                    pulados += 1
                    continue
                db.add(Blacklist(tipo=tipo_raw, valor=valor, motivo=motivo, fonte=fonte, ativo=True))
                inseridos += 1
        except Exception as exc:
            erros.append({"linha": i, "erro": str(exc)})

    db.commit()

    resultado = {"inseridos": inseridos, "pulados": pulados, "erros": len(erros), "detalhes_erros": erros[:50]}

    log_auditoria(
        db,
        acao=(
            f"Importou blacklist em lote ({arquivo.filename}): "
            f"{inseridos} inseridos, {pulados} já existentes, {len(erros)} com erro"
        ),
        usuario=usuario,
        request=request,
        tipo_entidade="blacklist",
        entidade_id=None,
        depois={"inseridos": inseridos, "pulados": pulados, "erros": len(erros)},
        risco="MEDIO",
    )
    db.commit()

    return resultado

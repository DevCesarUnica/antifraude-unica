"""
Router de corretores — CRUD completo + contatos + importação CSV + visão unificada Storm.
"""

import asyncio
import csv
import io
import uuid
from datetime import datetime, timezone
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, status, Request, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, joinedload
from sqlalchemy.exc import IntegrityError

from app.database import get_db
from app.models import Corretor, ContatoCorretor, ImportacaoCorretor
from app.schemas import (
    CorretorCreate, CorretorUpdate, CorretorOut,
    ContatoCreate, ContatoOut, ImportacaoOut, Mensagem,
)

router = APIRouter(prefix="/corretores", tags=["corretores"])


# ── Normalizadores ─────────────────────────────────────────────────────────────

def _normalize_interno(c: Corretor) -> dict:
    grupo_nome = c.grupo.nome if c.grupo else None
    return {
        "id": str(c.id),
        "codigo": c.codigo_externo or c.cpf,
        "nome": c.nome or "—",
        "email": c.email,
        "status": "ativo" if c.ativo else "inativo",
        "tipo": grupo_nome or "Corretor Interno",
        "loja": None,
        "privilegio": None,
        "origem": "interno",
        "criado_em": c.criado_em.isoformat() if c.criado_em else None,
    }


def _normalize_storm(raw: dict) -> dict:
    loja_sala = raw.get("loja_sala") or {}
    partes_loja = []
    if isinstance(loja_sala, dict):
        if loja_sala.get("gc"):
            partes_loja.append(f"GC: {loja_sala['gc']}")
        if loja_sala.get("gd"):
            partes_loja.append(f"GD: {loja_sala['gd']}")
    loja = " · ".join(partes_loja) or None

    priv_obj = raw.get("privilegio")
    privilegio = priv_obj.get("descricao") if isinstance(priv_obj, dict) else None

    return {
        "id": f"storm_{raw['id']}",
        "codigo": raw.get("usuario") or str(raw["id"]),
        "nome": raw.get("nome") or "—",
        "email": raw.get("email") or None,
        "status": "ativo" if raw.get("status") == 1 else "inativo",
        "tipo": privilegio or "Colaborador Storm",
        "loja": loja,
        "privilegio": privilegio,
        "origem": "storm",
        "criado_em": raw.get("data_cadastro"),
    }


# ── Endpoint unificado ────────────────────────────────────────────────────────

@router.get("/unificados")
async def listar_corretores_unificados(
    pagina: int = Query(1, ge=1),
    nome: str | None = None,
    codigo: str | None = None,      # busca por código/usuário Storm ou CPF/código_externo interno
    status_filtro: str | None = Query(None, alias="status"),
    origem: str | None = None,
    db: Session = Depends(get_db),
):
    """
    Lista unificada de corretores internos + colaboradores Storm.

    - origem=interno  → apenas banco local
    - origem=storm    → apenas Storm API
    - sem origem      → ambos (interno sempre incluso; Storm paginado)
    - codigo          → busca por usuario Storm ou CPF/codigo_externo interno
    """
    from app.services.storm import StormService, StormAPIError

    sync_em = datetime.now(timezone.utc).isoformat()
    items: list[dict] = []
    paginacao_storm: dict | None = None

    # ── 1. Corretores internos ───────────────────────────────────────────────
    if origem in (None, "interno"):
        q = db.query(Corretor).options(joinedload(Corretor.grupo))
        if nome:
            q = q.filter(Corretor.nome.ilike(f"%{nome}%"))
        if codigo:
            codigo_limpo = codigo.replace(".", "").replace("-", "").strip()
            q = q.filter(
                (Corretor.cpf == codigo_limpo) |
                (Corretor.codigo_externo.ilike(f"%{codigo}%"))
            )
        if status_filtro == "ativo":
            q = q.filter(Corretor.ativo == True)
        elif status_filtro == "inativo":
            q = q.filter(Corretor.ativo == False)
        internos = q.order_by(Corretor.nome.asc()).all()
        items.extend(_normalize_interno(c) for c in internos)

    # ── 2. Colaboradores Storm ───────────────────────────────────────────────
    if origem in (None, "storm"):
        try:
            async with StormService() as storm:
                status_usuario: str | None = None
                if status_filtro == "ativo":
                    status_usuario = "1"
                elif status_filtro == "inativo":
                    status_usuario = "0"

                res = await storm.get_colaboradores(
                    pagina=pagina,
                    usuario=codigo or None,     # código exato vai direto para a Storm
                    status_usuario=status_usuario,
                )

            colaboradores = res.get("data") or []
            for raw in colaboradores:
                norm = _normalize_storm(raw)
                # filtro de nome: Storm não suporta, aplica no cliente
                if nome and nome.lower() not in (norm["nome"] or "").lower():
                    continue
                items.append(norm)

            paginacao_storm = {
                "total": res.get("total", 0),
                "paginas": res.get("last_page", 1),
                "pagina_atual": res.get("current_page", pagina),
                "por_pagina": res.get("per_page", 50),
            }
        except StormAPIError:
            paginacao_storm = None

    return {
        "items": items,
        "paginacao_storm": paginacao_storm,
        "sync_em": sync_em,
    }


# ── Exportação Excel (assíncrona, com progresso) ─────────────────────────────
# Mesmos filtros do endpoint unificado, mas percorre TODAS as páginas da Storm
# (não só a página atual) para trazer a base completa.
#
# A busca completa na Storm pode levar minutos (rate limit de 20 req/min +
# retries), então roda em background: `iniciar` devolve um job_id na hora,
# o front consulta `status/{job_id}` periodicamente para mostrar % concluído,
# e baixa o arquivo em `download/{job_id}` quando `status == concluido`.
#
# Estado em memória — assume processo único de dev/produção (sem múltiplos
# workers). Jobs concluídos são removidos do dicionário após o download.

_COLUNAS_EXCEL_CORRETORES = ["Código", "Nome", "E-mail", "Status", "Tipo", "Origem", "Loja/Sala", "Privilégio", "Criado Em"]
_LARGURAS_EXCEL_CORRETORES = [16, 28, 28, 10, 22, 10, 20, 20, 18]

# Trava de segurança contra loop indefinido caso a Storm devolva `last_page`
# inconsistente entre chamadas.
_STORM_MAX_PAGINAS = 2000

_EXPORT_JOBS: dict[str, dict] = {}


def _montar_workbook_corretores(itens: list[dict], storm_indisponivel: bool) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Corretores"

    ws.append(_COLUNAS_EXCEL_CORRETORES)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    for col_idx in range(1, len(_COLUNAS_EXCEL_CORRETORES) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for item in itens:
        ws.append([
            item["codigo"],
            item["nome"],
            item["email"] or "",
            item["status"].upper(),
            item["tipo"] or "",
            item["origem"].upper(),
            item["loja"] or "",
            item["privilegio"] or "",
            item["criado_em"] or "",
        ])

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    for col_idx, largura in enumerate(_LARGURAS_EXCEL_CORRETORES, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = largura

    if storm_indisponivel:
        ws2 = wb.create_sheet("Avisos")
        ws2.append(["Colaboradores Storm indisponíveis no momento da exportação — lista abaixo inclui apenas corretores internos."])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


async def _executar_exportacao_corretores(
    job_id: str,
    nome: str | None,
    codigo: str | None,
    status_filtro: str | None,
    origem: str | None,
) -> None:
    from app.database import SessionLocal
    from app.services.storm import StormService, StormAPIError

    job = _EXPORT_JOBS[job_id]
    db = SessionLocal()
    try:
        itens: list[dict] = []

        if origem in (None, "interno"):
            q = db.query(Corretor).options(joinedload(Corretor.grupo))
            if nome:
                q = q.filter(Corretor.nome.ilike(f"%{nome}%"))
            if codigo:
                codigo_limpo = codigo.replace(".", "").replace("-", "").strip()
                q = q.filter(
                    (Corretor.cpf == codigo_limpo) |
                    (Corretor.codigo_externo.ilike(f"%{codigo}%"))
                )
            if status_filtro == "ativo":
                q = q.filter(Corretor.ativo == True)  # noqa: E712
            elif status_filtro == "inativo":
                q = q.filter(Corretor.ativo == False)  # noqa: E712
            internos = q.order_by(Corretor.nome.asc()).all()
            itens.extend(_normalize_interno(c) for c in internos)

        storm_indisponivel = False
        if origem in (None, "storm"):
            status_usuario: str | None = None
            if status_filtro == "ativo":
                status_usuario = "1"
            elif status_filtro == "inativo":
                status_usuario = "0"

            try:
                async with StormService() as storm:
                    pagina = 1
                    paginas_totais = 1
                    while pagina <= paginas_totais and pagina <= _STORM_MAX_PAGINAS:
                        res = await storm.get_colaboradores(
                            pagina=pagina,
                            usuario=codigo or None,
                            status_usuario=status_usuario,
                        )
                        for raw in res.get("data") or []:
                            norm = _normalize_storm(raw)
                            if nome and nome.lower() not in (norm["nome"] or "").lower():
                                continue
                            itens.append(norm)
                        paginas_totais = res.get("last_page", 1)
                        job["paginas_processadas"] = pagina
                        job["paginas_totais"] = paginas_totais
                        pagina += 1
            except StormAPIError:
                storm_indisponivel = True

        job["arquivo"] = _montar_workbook_corretores(itens, storm_indisponivel)
        job["nome_arquivo"] = f"corretores_{datetime.now(timezone.utc).strftime('%Y-%m-%d_%H-%M')}.xlsx"
        job["status"] = "concluido"
    except Exception as exc:
        job["status"] = "erro"
        job["erro"] = str(exc)
    finally:
        db.close()


@router.post("/exportar-excel/iniciar")
async def iniciar_exportacao_corretores(
    nome: str | None = None,
    codigo: str | None = None,
    status_filtro: str | None = Query(None, alias="status"),
    origem: str | None = None,
):
    job_id = str(uuid.uuid4())
    _EXPORT_JOBS[job_id] = {
        "status": "em_andamento",
        "paginas_processadas": 0,
        "paginas_totais": 1,
        "erro": None,
        "arquivo": None,
        "nome_arquivo": None,
    }
    asyncio.create_task(_executar_exportacao_corretores(job_id, nome, codigo, status_filtro, origem))
    return {"job_id": job_id}


@router.get("/exportar-excel/status/{job_id}")
def status_exportacao_corretores(job_id: str):
    job = _EXPORT_JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job de exportação não encontrado ou já expirado")
    total = max(job["paginas_totais"], 1)
    percentual = 100 if job["status"] == "concluido" else min(99, int(job["paginas_processadas"] / total * 100))
    return {
        "status": job["status"],
        "percentual": percentual,
        "paginas_processadas": job["paginas_processadas"],
        "paginas_totais": job["paginas_totais"],
        "erro": job["erro"],
    }


@router.get("/exportar-excel/download/{job_id}")
def download_exportacao_corretores(job_id: str):
    job = _EXPORT_JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job de exportação não encontrado ou já expirado")
    if job["status"] != "concluido":
        raise HTTPException(status_code=409, detail="Exportação ainda não concluída")

    buffer = io.BytesIO(job["arquivo"])
    nome_arquivo = job["nome_arquivo"]
    del _EXPORT_JOBS[job_id]

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome_arquivo}"'},
    )


# ── CRUD ──────────────────────────────────────────────────────────────────────

@router.get("/", response_model=list[CorretorOut])
def listar_corretores(
    nome: str | None = None,
    cpf: str | None = None,
    grupo_id: str | None = None,
    ativo: bool | None = None,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
):
    q = db.query(Corretor)
    if nome:
        q = q.filter(Corretor.nome.ilike(f"%{nome}%"))
    if cpf:
        q = q.filter(Corretor.cpf == cpf.replace(".", "").replace("-", ""))
    if grupo_id:
        q = q.filter(Corretor.grupo_id == grupo_id)
    if ativo is not None:
        q = q.filter(Corretor.ativo == ativo)
    return q.order_by(Corretor.nome.asc()).offset(skip).limit(limit).all()


@router.post("/", response_model=CorretorOut, status_code=status.HTTP_201_CREATED)
def criar_corretor(body: CorretorCreate, db: Session = Depends(get_db)):
    corretor = Corretor(**body.model_dump())
    db.add(corretor)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="CPF ou código externo já cadastrado")
    db.refresh(corretor)
    return corretor


@router.get("/importacoes/historico", response_model=list[ImportacaoOut])
def historico_importacoes(skip: int = 0, limit: int = 50, db: Session = Depends(get_db)):
    return db.query(ImportacaoCorretor).order_by(
        ImportacaoCorretor.criado_em.desc()
    ).offset(skip).limit(limit).all()


@router.get("/{corretor_id}", response_model=CorretorOut)
def obter_corretor(corretor_id: str, db: Session = Depends(get_db)):
    return _get_ou_404(db, corretor_id)


@router.patch("/{corretor_id}", response_model=CorretorOut)
def atualizar_corretor(corretor_id: str, body: CorretorUpdate, db: Session = Depends(get_db)):
    c = _get_ou_404(db, corretor_id)
    for campo, valor in body.model_dump(exclude_none=True).items():
        setattr(c, campo, valor)
    db.commit()
    db.refresh(c)
    return c


@router.delete("/{corretor_id}", response_model=Mensagem)
def desativar_corretor(corretor_id: str, db: Session = Depends(get_db)):
    c = _get_ou_404(db, corretor_id)
    c.ativo = False
    db.commit()
    return Mensagem(mensagem="Corretor desativado com sucesso")


# ── Contatos ──────────────────────────────────────────────────────────────────

@router.get("/{corretor_id}/contatos", response_model=list[ContatoOut])
def listar_contatos(corretor_id: str, db: Session = Depends(get_db)):
    _get_ou_404(db, corretor_id)
    return db.query(ContatoCorretor).filter(
        ContatoCorretor.corretor_id == corretor_id,
        ContatoCorretor.ativo == True,
    ).all()


@router.post("/{corretor_id}/contatos", response_model=ContatoOut, status_code=status.HTTP_201_CREATED)
def adicionar_contato(corretor_id: str, body: ContatoCreate, db: Session = Depends(get_db)):
    _get_ou_404(db, corretor_id)
    if body.principal:
        # Remove flag principal dos outros do mesmo tipo
        db.query(ContatoCorretor).filter(
            ContatoCorretor.corretor_id == corretor_id,
            ContatoCorretor.tipo == body.tipo,
        ).update({"principal": False})
    contato = ContatoCorretor(corretor_id=corretor_id, **body.model_dump())
    db.add(contato)
    db.commit()
    db.refresh(contato)
    return contato


@router.delete("/{corretor_id}/contatos/{contato_id}", response_model=Mensagem)
def remover_contato(corretor_id: str, contato_id: str, db: Session = Depends(get_db)):
    contato = db.query(ContatoCorretor).filter(
        ContatoCorretor.id == contato_id,
        ContatoCorretor.corretor_id == corretor_id,
    ).first()
    if not contato:
        raise HTTPException(status_code=404, detail="Contato não encontrado")
    contato.ativo = False
    db.commit()
    return Mensagem(mensagem="Contato removido")


# ── Importação CSV ────────────────────────────────────────────────────────────

@router.post("/importar", response_model=ImportacaoOut, status_code=status.HTTP_201_CREATED)
async def importar_corretores(
    request: Request,
    arquivo: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """
    Importa corretores via CSV.
    Colunas esperadas: nome, cpf, email, telefone, codigo_externo, grupo_id
    """
    conteudo = await arquivo.read()
    try:
        texto = conteudo.decode("utf-8-sig")
    except UnicodeDecodeError:
        texto = conteudo.decode("latin-1")

    importacao = ImportacaoCorretor(
        arquivo_nome=arquivo.filename,
        status="PROCESSANDO",
        criado_por=request.headers.get("x-usuario"),
    )
    db.add(importacao)
    db.flush()

    leitor = csv.DictReader(io.StringIO(texto))
    erros = []
    sucesso = 0
    total = 0

    for i, linha in enumerate(leitor, start=2):
        total += 1
        try:
            cpf = linha.get("cpf", "").replace(".", "").replace("-", "").strip()
            if not cpf:
                raise ValueError("CPF obrigatório")
            nome = linha.get("nome", "").strip()
            if not nome:
                raise ValueError("Nome obrigatório")

            existente = db.query(Corretor).filter(Corretor.cpf == cpf).first()
            if existente:
                existente.nome = nome
                existente.email = linha.get("email", "").strip() or existente.email
                existente.telefone = linha.get("telefone", "").strip() or existente.telefone
            else:
                c = Corretor(
                    nome=nome,
                    cpf=cpf,
                    email=linha.get("email", "").strip() or None,
                    telefone=linha.get("telefone", "").strip() or None,
                    codigo_externo=linha.get("codigo_externo", "").strip() or None,
                    grupo_id=linha.get("grupo_id", "").strip() or None,
                )
                db.add(c)
            sucesso += 1
        except Exception as exc:
            erros.append({"linha": i, "erro": str(exc), "dados": dict(linha)})

    importacao.total_linhas = total
    importacao.sucesso = sucesso
    importacao.erro = len(erros)
    importacao.log_erros = erros if erros else None
    importacao.status = "CONCLUIDO" if not erros or sucesso > 0 else "ERRO"

    from datetime import datetime
    importacao.concluido_em = datetime.utcnow()
    db.commit()
    db.refresh(importacao)
    return importacao


# ── Helper ────────────────────────────────────────────────────────────────────

def _get_ou_404(db: Session, corretor_id: str) -> Corretor:
    c = db.query(Corretor).filter(Corretor.id == corretor_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Corretor não encontrado")
    return c

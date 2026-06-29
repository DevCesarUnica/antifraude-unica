"""
Router de logs — acesso HTTP e auditoria completa de ações de usuários.
"""

import io
import json
from datetime import datetime, timedelta, timezone
from zoneinfo import ZoneInfo
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy import func, extract, or_
from sqlalchemy.orm import Session

BRT = ZoneInfo("America/Sao_Paulo")

from app.database import get_db
from app.models import LogAcesso, LogAuditoria, Usuario
from app.schemas import LogAcessoOut, LogAuditoriaOut
from app.routers.auth import verificar_token

router = APIRouter(prefix="/logs", tags=["logs"])


# ── Logs de acesso HTTP ───────────────────────────────────────────────────────

@router.get("/acesso", response_model=list[LogAcessoOut])
def listar_logs(
    usuario_id: str | None = None,
    metodo: str | None = None,
    endpoint: str | None = None,
    status_code: int | None = None,
    data_inicio: datetime | None = None,
    data_fim: datetime | None = None,
    skip: int = 0,
    limit: int = 200,
    db: Session = Depends(get_db),
):
    q = (
        db.query(LogAcesso, Usuario.nome, Usuario.perfil)
        .outerjoin(Usuario, LogAcesso.usuario_id == Usuario.id)
    )
    if usuario_id:
        q = q.filter(LogAcesso.usuario_id == usuario_id)
    if metodo:
        q = q.filter(LogAcesso.metodo == metodo.upper())
    if endpoint:
        q = q.filter(LogAcesso.endpoint.ilike(f"%{endpoint}%"))
    if status_code:
        q = q.filter(LogAcesso.status_code == status_code)
    if data_inicio:
        q = q.filter(LogAcesso.timestamp >= data_inicio)
    if data_fim:
        q = q.filter(LogAcesso.timestamp <= data_fim)

    rows = q.order_by(LogAcesso.timestamp.desc()).offset(skip).limit(limit).all()
    return [
        {
            "id": log.id,
            "usuario_id": log.usuario_id,
            "username": log.username,
            "nome": nome,
            "perfil": str(perfil) if perfil else None,
            "metodo": log.metodo,
            "endpoint": log.endpoint,
            "ip": log.ip,
            "status_code": log.status_code,
            "duracao_ms": log.duracao_ms,
            "timestamp": log.timestamp,
        }
        for log, nome, perfil in rows
    ]


@router.get("/acesso/resumo")
def resumo_logs(db: Session = Depends(get_db)):
    corte = datetime.now(timezone.utc) - timedelta(hours=24)
    resultados = (
        db.query(LogAcesso.status_code, func.count(LogAcesso.id).label("total"))
        .filter(LogAcesso.timestamp >= corte)
        .group_by(LogAcesso.status_code)
        .all()
    )
    return {"por_status": {str(r.status_code): r.total for r in resultados}}


# ── Auditoria de ações de usuários ────────────────────────────────────────────

@router.get("/auditoria", response_model=list[LogAuditoriaOut])
def listar_auditoria(
    usuario: str | None = None,
    acao: str | None = None,
    tipo_entidade: str | None = None,
    risco: str | None = None,
    sucesso: bool | None = None,
    data_inicio: datetime | None = None,
    data_fim: datetime | None = None,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    _: Usuario = Depends(verificar_token),
):
    q = db.query(LogAuditoria)
    if usuario:
        q = q.filter(
            LogAuditoria.username.ilike(f"%{usuario}%") |
            LogAuditoria.nome.ilike(f"%{usuario}%")
        )
    if acao:
        q = q.filter(LogAuditoria.acao.ilike(f"%{acao}%"))
    if tipo_entidade:
        q = q.filter(LogAuditoria.tipo_entidade == tipo_entidade)
    if risco:
        q = q.filter(LogAuditoria.risco == risco.upper())
    if sucesso is not None:
        q = q.filter(LogAuditoria.sucesso == sucesso)
    if data_inicio:
        q = q.filter(LogAuditoria.criado_em >= data_inicio)
    if data_fim:
        q = q.filter(LogAuditoria.criado_em <= data_fim)
    return q.order_by(LogAuditoria.criado_em.desc()).offset(skip).limit(limit).all()


@router.get("/auditoria/exportar")
def exportar_auditoria_excel(
    usuario: str | None = None,
    acao: str | None = None,
    tipo_entidade: str | None = None,
    risco: str | None = None,
    sucesso: bool | None = None,
    data_inicio: datetime | None = None,
    data_fim: datetime | None = None,
    db: Session = Depends(get_db),
    _: Usuario = Depends(verificar_token),
):
    """Exporta os logs de auditoria filtrados em formato Excel (.xlsx)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── Busca com os mesmos filtros do endpoint de listagem ──────────────────
    q = db.query(LogAuditoria)
    if usuario:
        q = q.filter(
            LogAuditoria.username.ilike(f"%{usuario}%") |
            LogAuditoria.nome.ilike(f"%{usuario}%")
        )
    if acao:
        q = q.filter(LogAuditoria.acao.ilike(f"%{acao}%"))
    if tipo_entidade:
        q = q.filter(LogAuditoria.tipo_entidade == tipo_entidade)
    if risco:
        q = q.filter(LogAuditoria.risco == risco.upper())
    if sucesso is not None:
        q = q.filter(LogAuditoria.sucesso == sucesso)
    if data_inicio:
        q = q.filter(LogAuditoria.criado_em >= data_inicio)
    if data_fim:
        q = q.filter(LogAuditoria.criado_em <= data_fim)

    registros = q.order_by(LogAuditoria.criado_em.desc()).limit(10000).all()

    # ── Monta o workbook ─────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Auditoria"

    # Cores
    COR_HEADER   = "1C1C1E"
    COR_ALTO     = "FEE2E2"
    COR_MEDIO    = "FEF9C3"
    COR_BAIXO    = "DCFCE7"
    COR_ERRO     = "FEE2E2"
    COR_SUCESSO  = "DCFCE7"

    borda = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    # Cabeçalhos
    cabecalhos = [
        "Data / Hora",
        "Usuário",
        "Username",
        "Perfil",
        "Ação",
        "Tipo Entidade",
        "ID Entidade",
        "Risco",
        "Sucesso",
        "IP",
        "Navegador",
        "Origem",
        "Erro",
        "Antes (JSON)",
        "Depois (JSON)",
    ]

    # Linha de título
    ws.merge_cells("A1:O1")
    titulo = ws["A1"]
    titulo.value = f"Auditoria de Ações — exportado em {datetime.now(timezone.utc).astimezone(BRT).strftime('%d/%m/%Y %H:%M')} (Horário de Brasília)"
    titulo.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    titulo.fill = PatternFill("solid", fgColor=COR_HEADER)
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Linha de cabeçalhos
    for col_idx, nome in enumerate(cabecalhos, start=1):
        cell = ws.cell(row=2, column=col_idx, value=nome)
        cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="374151")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borda
    ws.row_dimensions[2].height = 22

    # Dados
    for row_idx, r in enumerate(registros, start=3):
        data_hora = r.criado_em.astimezone(BRT).strftime("%d/%m/%Y %H:%M:%S") if r.criado_em else ""
        valores = [
            data_hora,
            r.nome or "",
            r.username or "",
            r.perfil or "",
            r.acao or "",
            r.tipo_entidade or "",
            r.entidade_id or "",
            r.risco or "",
            "Sim" if r.sucesso else "Não",
            r.ip or "",
            (r.user_agent or "")[:100],
            r.origem or "",
            r.erro or "",
            json.dumps(r.antes, ensure_ascii=False) if r.antes else "",
            json.dumps(r.depois, ensure_ascii=False) if r.depois else "",
        ]

        # Cor de fundo por risco
        if r.risco == "ALTO":
            bg = COR_ALTO
        elif r.risco == "MEDIO":
            bg = COR_MEDIO
        else:
            bg = COR_BAIXO

        for col_idx, valor in enumerate(valores, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.font = Font(name="Calibri", size=9)
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border = borda
            # Coluna Sucesso: verde/vermelho independente do risco
            if col_idx == 9:
                cell.fill = PatternFill("solid", fgColor=COR_SUCESSO if r.sucesso else COR_ERRO)
            else:
                cell.fill = PatternFill("solid", fgColor=bg)

    # Larguras das colunas
    larguras = [18, 22, 16, 12, 45, 14, 22, 8, 8, 14, 30, 8, 30, 35, 35]
    for i, larg in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = larg

    # Congela cabeçalhos
    ws.freeze_panes = "A3"

    # Auto-filtro
    ws.auto_filter.ref = f"A2:{get_column_letter(len(cabecalhos))}2"

    # ── Gera o arquivo em memória e retorna ──────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nome_arquivo = f"auditoria_{datetime.now(timezone.utc).astimezone(BRT).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome_arquivo}"'},
    )


@router.get("/auditoria/resumo")
def resumo_auditoria(
    db: Session = Depends(get_db),
    _: Usuario = Depends(verificar_token),
):
    corte = datetime.now(timezone.utc) - timedelta(hours=24)
    total = db.query(func.count(LogAuditoria.id)).filter(LogAuditoria.criado_em >= corte).scalar() or 0
    alto_risco = db.query(func.count(LogAuditoria.id)).filter(
        LogAuditoria.criado_em >= corte, LogAuditoria.risco == "ALTO"
    ).scalar() or 0
    usuarios_ativos = db.query(func.count(func.distinct(LogAuditoria.usuario_id))).filter(
        LogAuditoria.criado_em >= corte, LogAuditoria.usuario_id.isnot(None)
    ).scalar() or 0
    erros = db.query(func.count(LogAuditoria.id)).filter(
        LogAuditoria.criado_em >= corte, LogAuditoria.sucesso == False
    ).scalar() or 0

    taxa_sucesso = round(((total - erros) / total * 100) if total > 0 else 100.0, 1)
    return {
        "total_acoes": total,
        "alto_risco": alto_risco,
        "usuarios_ativos": usuarios_ativos,
        "taxa_sucesso": taxa_sucesso,
    }


@router.get("/suspeitos")
def logs_suspeitos(
    db: Session = Depends(get_db),
    _: Usuario = Depends(verificar_token),
):
    """
    Detecta atividades suspeitas nas últimas 24h:
    - Concentração de ações de alto risco por usuário em 1h (≥3)
    - Múltiplos logins falhados do mesmo IP em 1h (≥3)
    - Ações de risco médio/alto fora do horário comercial (antes 10h UTC = antes 7h BRT)
    """
    corte_24h = datetime.now(timezone.utc) - timedelta(hours=24)
    corte_1h = datetime.now(timezone.utc) - timedelta(hours=1)
    suspeitos = []

    # 1. Concentração de alto risco por usuário em 1h
    alto_risco_rows = (
        db.query(
            LogAuditoria.usuario_id,
            LogAuditoria.username,
            LogAuditoria.nome,
            func.count(LogAuditoria.id).label("total"),
        )
        .filter(LogAuditoria.criado_em >= corte_1h, LogAuditoria.risco == "ALTO")
        .group_by(LogAuditoria.usuario_id, LogAuditoria.username, LogAuditoria.nome)
        .having(func.count(LogAuditoria.id) >= 3)
        .all()
    )
    for r in alto_risco_rows:
        suspeitos.append({
            "tipo": "concentracao_risco",
            "nivel": "ALTO",
            "descricao": f"{r.nome or r.username} realizou {r.total} ações de alto risco na última hora",
            "usuario_id": r.usuario_id,
            "username": r.username,
            "nome": r.nome,
        })

    # 2. Múltiplos logins falhados do mesmo IP em 1h
    logins_falhos = (
        db.query(LogAuditoria.ip, func.count(LogAuditoria.id).label("total"))
        .filter(
            LogAuditoria.criado_em >= corte_1h,
            LogAuditoria.acao.like("Tentativa de login%"),
            LogAuditoria.sucesso == False,
        )
        .group_by(LogAuditoria.ip)
        .having(func.count(LogAuditoria.id) >= 3)
        .all()
    )
    for r in logins_falhos:
        suspeitos.append({
            "tipo": "tentativas_login",
            "nivel": "ALTO",
            "descricao": f"IP {r.ip} tentou {r.total} logins falhados na última hora",
            "ip": r.ip,
        })

    # 3. Ações de risco médio/alto fora do horário comercial (07h–22h BRT)
    hora_brt = extract("hour", func.timezone("America/Sao_Paulo", LogAuditoria.criado_em))
    fora_horario = (
        db.query(LogAuditoria)
        .filter(
            LogAuditoria.criado_em >= corte_24h,
            LogAuditoria.risco.in_(["MEDIO", "ALTO"]),
            or_(hora_brt < 7, hora_brt >= 22),
        )
        .order_by(LogAuditoria.criado_em.desc())
        .limit(10)
        .all()
    )
    for r in fora_horario:
        criado_brt = r.criado_em.astimezone(BRT).strftime("%d/%m/%Y %H:%M:%S")
        suspeitos.append({
            "tipo": "fora_horario",
            "nivel": "MEDIO",
            "descricao": f"{r.nome or r.username or 'Sistema'} realizou '{r.acao}' fora do horário comercial ({criado_brt} BRT)",
            "log_id": r.id,
            "usuario_id": r.usuario_id,
            "username": r.username,
            "nome": r.nome,
            "acao": r.acao,
            "criado_em": r.criado_em.astimezone(BRT).isoformat(),
        })

    return {"suspeitos": suspeitos, "total": len(suspeitos)}
